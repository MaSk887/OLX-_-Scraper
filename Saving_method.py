import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def save_to_excel(all_results, AREA):
    if not all_results:
        print("no results")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "نتائج البحث"

    headers = ["المنطقة", "العنوان", "السعر", "التاريخ","غرف","حمام","مساحة","سعر المتر", "مالك", "الرابط"]

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_num, result in enumerate(all_results, 2):
        cell1 =ws.cell(row=row_num, column=1, value=result["المنطقة"])
        cell2 =ws.cell(row=row_num, column=2, value=result["العنوان"])

        cell3 = ws.cell(row=row_num, column=3, value=result["السعر"])
        cell3.number_format = '#,##0 "EGP"'
        cell8 = ws.cell(row=row_num, column=4, value=result["التاريخ"])


        cell4 = ws.cell(row=row_num, column=5, value=result["غرف"])
        cell5 = ws.cell(row=row_num, column=6, value=result["حمام"])
        cell6 = ws.cell(row=row_num, column=7, value=result["مساحة"])
        cell7 = ws.cell(row=row_num, column=8, value=result["سعر المتر"])

        cell9 = ws.cell(row=row_num, column=9, value=str(result["مالك"]))
        cell10 = ws.cell(row=row_num, column=10, value=result["الرابط"])

        for cell in [cell1, cell2, cell3, cell4, cell5, cell6, cell7, cell8, cell9]:
            cell.alignment = Alignment(horizontal="center", vertical="center")



    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 90
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 1000


    os.makedirs("results", exist_ok=True)


    ws.auto_filter.ref = f"A1:I{len(all_results) +1}"

    filename = os.path.join("results", f"all_Results_File_({AREA}).xlsx")

    try:
        wb.save(filename)
        print(f"\n{filename} saved")
        print(f"num of results : {len(all_results)}")
    except Exception as e:
        print(f"error in save: {e}")


